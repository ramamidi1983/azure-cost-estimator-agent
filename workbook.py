"""Generate a client-ready Excel workbook modeled on the reference RFP / Goodyear cost sheets.

Layout:
  * Summary        - Goodyear-style "Cost Area | Monthly ($) | Annual ($)" roll-up.
  * Detail         - RFP-style per-host inventory (Host Name, Env, OS, Server Type,
                     vCPU, Memory, Azure Type, Monthly, Annual, Storage).
  * Modernization  - optional per-app path comparison.
  * Mapping & Notes- SKU->vCPU/RAM mapping used + pricing assumptions.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLK = Font(name="Arial", size=10, color="000000")
BLKB = Font(name="Arial", size=10, bold=True, color="000000")
WHT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
TITLE = Font(name="Arial", size=14, bold=True, color="1F4E78")
SUBT = Font(name="Arial", size=9, italic=True, color="595959")
NOTE = Font(name="Arial", size=9, italic=True, color="C00000")
HDR = PatternFill("solid", fgColor="1F4E78")
SUB = PatternFill("solid", fgColor="D9E1F2")
TOTF = PatternFill("solid", fgColor="BDD7EE")
CUR0 = '$#,##0;($#,##0);"-"'
NUM0 = '#,##0;(#,##0);"-"'
NUM1 = '#,##0.0;(#,##0.0);"-"'
thin = Side(style="thin", color="BFBFBF")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR = Alignment(horizontal="center")
LFT = Alignment(horizontal="left")


def _hdr(ws, row, labels, start=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=l)
        c.font = WHT; c.fill = HDR; c.alignment = CTR; c.border = BORD


def _title(ws, title, subtitle, note=None):
    ws.sheet_view.showGridLines = False
    a = ws["A1"]; a.value = title; a.font = TITLE
    b = ws["A2"]; b.value = subtitle; b.font = SUBT
    if note:
        c = ws["A3"]; c.value = note; c.font = NOTE


def _summary(wb, summ, params):
    s = wb.active; s.title = "Summary"
    _title(s, "Azure Target Hosting - Cost Summary",
           f"Region: {params['region']}  |  Term: {params['term_label']}  |  "
           f"AHB: {'Yes' if params['ahb'] else 'No'}  |  Resiliency: {'Yes' if params['resiliency'] else 'No'}",
           f"Live Azure Retail Prices  |  Priced {params['generated']}  |  "
           "Steady-state hosting only - excludes one-time migration / professional services.")
    _hdr(s, 5, ["Cost Area", "Monthly ($)", "Annual ($)"])
    r = 6
    for _, row in summ.iterrows():
        is_total = str(row["area"]).upper() == "TOTAL"
        c1 = s.cell(r, 1, row["area"]); c1.border = BORD
        c2 = s.cell(r, 2, round(float(row["monthly"]), 2)); c2.number_format = CUR0; c2.border = BORD
        c3 = s.cell(r, 3, f"=B{r}*12"); c3.number_format = CUR0; c3.border = BORD
        if is_total:
            for c in (c1, c2, c3): c.font = WHT; c.fill = HDR
        else:
            c1.font = BLK; c2.font = BLK; c3.font = BLK
        r += 1
    for col, w in zip("ABC", [42, 16, 16]):
        s.column_dimensions[col].width = w


DETAIL_COLS = [
    ("name", "Host Name", 26, None),
    ("environment", "Env", 12, None),
    ("os", "OS", 12, None),
    ("role", "Role / Server Type", 20, None),
    ("disposition", "Disposition", 16, None),
    ("model", "Azure Service", 24, None),
    ("component", "Component", 12, None),
    ("sku", "Azure Type (SKU)", 22, None),
    ("vcpu", "vCPU", 8, NUM0),
    ("memory_gb", "Memory (GB)", 12, NUM1),
    ("storage_gb", "Storage (GB)", 12, NUM0),
    ("quantity", "Qty", 7, NUM0),
    ("monthly", "Monthly ($)", 14, CUR0),
]


def _detail(wb, lines):
    d = wb.create_sheet("Detail")
    _title(d, "Host Inventory - Azure Target & Monthly Cost",
           "One row per priced workload. Retain / Retire servers stay on-prem and are excluded.")
    keys = [k for k, *_ in DETAIL_COLS if k in lines.columns] if lines is not None and not lines.empty else []
    if not keys:
        d.cell(5, 1, "No priced workloads (all servers Retain/Retire or inventory empty).").font = BLK
        return
    labels = [lbl for k, lbl, *_ in DETAIL_COLS if k in keys] + ["Annual ($)"]
    _hdr(d, 5, labels)
    mcol = len(keys)
    acol = mcol + 1
    sort_cols = [c for c in ["environment", "model", "monthly"] if c in lines.columns]
    if sort_cols:
        asc = [True if c != "monthly" else False for c in sort_cols]
        df = lines.sort_values(sort_cols, ascending=asc)
    else:
        df = lines
    r = 6
    for _, row in df.iterrows():
        for j, k in enumerate(keys, start=1):
            fmt = next(f for kk, _l, _w, f in DETAIL_COLS if kk == k)
            v = row.get(k)
            if k == "monthly":
                v = round(float(v or 0), 2)
            cell = d.cell(r, j, v); cell.border = BORD; cell.font = BLK
            if fmt:
                cell.number_format = fmt
        ac = d.cell(r, acol, f"={get_column_letter(mcol)}{r}*12")
        ac.number_format = CUR0; ac.border = BORD; ac.font = BLK
        r += 1
    t1 = d.cell(r, 1, "TOTAL"); t1.font = WHT; t1.fill = HDR; t1.border = BORD
    for j in range(2, mcol):
        cc = d.cell(r, j); cc.fill = TOTF; cc.border = BORD
    ml = get_column_letter(mcol); al = get_column_letter(acol)
    mt = d.cell(r, mcol, f"=SUM({ml}6:{ml}{r-1})"); mt.number_format = CUR0; mt.font = WHT; mt.fill = HDR; mt.border = BORD
    at = d.cell(r, acol, f"=SUM({al}6:{al}{r-1})"); at.number_format = CUR0; at.font = WHT; at.fill = HDR; at.border = BORD
    d.freeze_panes = "A6"
    d.auto_filter.ref = f"A5:{al}5"
    widths = [w for k, _l, w, _f in DETAIL_COLS if k in keys] + [14]
    for i, w in enumerate(widths, start=1):
        d.column_dimensions[get_column_letter(i)].width = w


def _modernization(wb, modern):
    if modern is None or modern.empty:
        return
    mo = wb.create_sheet("Modernization")
    _title(mo, "Modernization Path Comparison",
           "Per app workload, $/month. Rehost vs Replatform vs Containerize vs Modernize.")
    headers = list(modern.columns)
    _hdr(mo, 5, [h.replace("_", " ").title() for h in headers])
    r = 6
    for _, row in modern.iterrows():
        is_total = str(row[headers[0]]).upper() == "TOTAL"
        for j, c in enumerate(headers, start=1):
            cell = mo.cell(r, j, row[c]); cell.border = BORD
            if j == 1:
                cell.font = BLKB if is_total else BLK
            else:
                cell.number_format = CUR0
                cell.font = BLKB if is_total else BLK
                if is_total:
                    cell.fill = SUB
        r += 1
    mo.column_dimensions["A"].width = 26
    for i in range(2, len(headers) + 1):
        mo.column_dimensions[get_column_letter(i)].width = 22


def _mapping(wb, lines, params):
    m = wb.create_sheet("Mapping & Notes")
    _title(m, "SKU Mapping, Pricing Sources & Assumptions", "")
    m.cell(4, 1, "Assumption").font = WHT; m.cell(4, 1).fill = HDR; m.cell(4, 1).border = BORD
    m.cell(4, 2, "Value").font = WHT; m.cell(4, 2).fill = HDR; m.cell(4, 2).border = BORD
    assumptions = [
        ("Region", params["region"]),
        ("Term / commitment", params["term_label"]),
        ("Azure Hybrid Benefit", "Yes" if params["ahb"] else "No"),
        ("Resiliency add-in", "Yes" if params["resiliency"] else "No"),
        ("Pricing source", "Azure Retail Prices API (live list prices, USD)"),
        ("Priced on", params["generated"]),
        ("Scope", "Steady-state hosting only; excludes migration/professional services"),
        ("Excluded", "Retain / Retire dispositions stay on-prem (not priced)"),
    ]
    r = 5
    for k, v in assumptions:
        a = m.cell(r, 1, k); a.font = BLKB; a.border = BORD
        b = m.cell(r, 2, v); b.font = BLK; b.border = BORD; b.alignment = LFT
        r += 1
    r += 1
    m.cell(r, 1, "Azure Type (SKU)").font = WHT; m.cell(r, 1).fill = HDR; m.cell(r, 1).border = BORD
    m.cell(r, 2, "vCPU / RAM (GB)").font = WHT; m.cell(r, 2).fill = HDR; m.cell(r, 2).border = BORD
    m.cell(r, 3, "Azure Service").font = WHT; m.cell(r, 3).fill = HDR; m.cell(r, 3).border = BORD
    r += 1
    if lines is not None and not lines.empty and "sku" in lines.columns:
        seen = set()
        cols = lines.columns
        for _, row in lines.iterrows():
            sku = str(row.get("sku", "") or "")
            if not sku or sku in seen:
                continue
            seen.add(sku)
            vc = row.get("vcpu", "") if "vcpu" in cols else ""
            mem = row.get("memory_gb", "") if "memory_gb" in cols else ""
            spec = f"{int(vc) if vc else '-'} / {int(mem) if mem else '-'}" if (vc or mem) else "-"
            m.cell(r, 1, sku).font = BLK; m.cell(r, 1).border = BORD
            m.cell(r, 2, spec).font = BLK; m.cell(r, 2).border = BORD; m.cell(r, 2).alignment = CTR
            m.cell(r, 3, str(row.get("model", "") or "")).font = BLK; m.cell(r, 3).border = BORD
            r += 1
    for col, w in zip("ABC", [34, 26, 26]):
        m.column_dimensions[col].width = w


def build(lines, summ, params, out_path, modern=None):
    wb = Workbook()
    _summary(wb, summ, params)
    _detail(wb, lines)
    _modernization(wb, modern)
    _mapping(wb, lines, params)
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path
